#!/usr/bin/env python3
"""
엑셀 파일 DB 적재 (openpyxl read_only + pymysql batch INSERT)
- 기존 데이터 유지 (DELETE/TRUNCATE 없음!)
- /BIC/ 접두사 자동 제거
- autocommit=True (대용량 트랜잭션 방지)
- BATCH_SIZE 300 (메모리 절약)
- 출력: JSON (stdout) - 마지막 줄만 JSON
"""
import sys
import json
import time
import gc
import pymysql
from openpyxl import load_workbook

BATCH_SIZE = 300

def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "사용법: xlsx_load.py <파일경로> <매핑JSON>"}))
        sys.exit(1)

    xlsx_path = sys.argv[1]
    mapped_cols = json.loads(sys.argv[2])

    try:
        total_start = time.time()

        conn = pymysql.connect(
            host='localhost', user='company', password='company1234!',
            database='company_board', charset='utf8mb4',
            autocommit=True
        )
        cursor = conn.cursor()

        # 기존 행수 확인
        cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
        before_count = cursor.fetchone()[0]
        print(f"[적재] 기존 DB 행수: {before_count:,}행 (유지됨)", file=sys.stderr)

        # DB 컬럼 세트 확인
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA='company_board' AND TABLE_NAME='bw_profitability_data'
            AND COLUMN_NAME != 'SEQ'
            ORDER BY ORDINAL_POSITION
        """)
        db_col_set = set(row[0] for row in cursor.fetchall())

        # 엑셀 읽기
        print(f"[적재] 엑셀 파일 열기...", file=sys.stderr)
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        print(f"[적재] 시트: {ws.title}, max_row={ws.max_row}", file=sys.stderr)

        row_iter = ws.iter_rows(values_only=True)
        next(row_iter)  # 1행 스킵 (한글 주석)

        # 2행: 영문 컬럼명 → /BIC/ 제거 후 DB 매핑
        excel_cols = [str(c or '').strip() for c in next(row_iter)]
        col_map = []
        for i, col in enumerate(excel_cols):
            clean_col = col.replace('/BIC/', '') if col.startswith('/BIC/') else col
            if clean_col and clean_col in db_col_set:
                col_map.append((i, clean_col))
        print(f"[적재] 매핑: {len(col_map)}/{len(excel_cols)}개 컬럼", file=sys.stderr)

        col_list = ', '.join(f'`{c[1]}`' for c in col_map)
        placeholders = ', '.join(['%s'] * len(col_map))
        insert_sql = f"INSERT INTO bw_profitability_data ({col_list}) VALUES ({placeholders})"

        # 적재 시작
        print(f"[적재] 데이터 INSERT 시작...", file=sys.stderr)
        insert_start = time.time()
        inserted = 0
        errors = 0
        error_details = []
        total_rows = 0
        batch = []

        for row in row_iter:
            row_list = list(row)
            has_value = False
            for idx, _ in col_map:
                v = row_list[idx] if idx < len(row_list) else None
                if v is not None and str(v).strip() != '':
                    has_value = True
                    break
            if not has_value:
                continue

            total_rows += 1
            values = []
            for idx, col_name in col_map:
                v = row_list[idx] if idx < len(row_list) else None
                if v is None or (isinstance(v, str) and v.strip() == ''):
                    values.append(None)
                else:
                    values.append(v)
            batch.append(tuple(values))

            if len(batch) >= BATCH_SIZE:
                try:
                    cursor.executemany(insert_sql, batch)
                    inserted += len(batch)
                except Exception as e:
                    for j, rv in enumerate(batch):
                        try:
                            cursor.execute(insert_sql, rv)
                            inserted += 1
                        except Exception as re:
                            errors += 1
                            if len(error_details) < 50:
                                error_details.append({
                                    "row": total_rows - len(batch) + j + 3,
                                    "error": str(re)[:200]
                                })
                batch = []
                if total_rows % 5000 < BATCH_SIZE:
                    gc.collect()
                    elapsed = time.time() - insert_start
                    rate = inserted / elapsed if elapsed > 0 else 0
                    print(f"  [진행] {total_rows:,}행 / {inserted:,} INSERT ({rate:.0f}행/초)", file=sys.stderr)

        # 남은 배치
        if batch:
            try:
                cursor.executemany(insert_sql, batch)
                inserted += len(batch)
            except:
                for j, rv in enumerate(batch):
                    try:
                        cursor.execute(insert_sql, rv)
                        inserted += 1
                    except Exception as re:
                        errors += 1
                        if len(error_details) < 50:
                            error_details.append({
                                "row": total_rows - len(batch) + j + 3,
                                "error": str(re)[:200]
                            })

        wb.close()
        gc.collect()

        # 최종 확인
        cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
        after_count = cursor.fetchone()[0]

        cursor.execute("SELECT CALMONTH, COUNT(*) AS cnt FROM bw_profitability_data GROUP BY CALMONTH ORDER BY CALMONTH")
        calmonth_dist = [{"calmonth": m, "count": c} for m, c in cursor.fetchall()]

        elapsed_total = time.time() - total_start
        print(f"[적재] 완료: {inserted:,}/{total_rows:,}행 INSERT, {errors}건 실패, {elapsed_total:.1f}초", file=sys.stderr)

        cursor.close()
        conn.close()

        # 결과 JSON (stdout 마지막 줄)
        result = {
            "success": True,
            "beforeRows": before_count,
            "totalExcelRows": total_rows,
            "insertedRows": inserted,
            "failedRows": errors,
            "totalDbRows": after_count,
            "addedRows": after_count - before_count,
            "elapsedSec": round(elapsed_total, 1),
            "mappedColumns": [c[1] for c in col_map],
            "calmonthDist": calmonth_dist,
            "errors": error_details,
        }
        print(json.dumps(result, ensure_ascii=False))

    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)

if __name__ == '__main__':
    main()
