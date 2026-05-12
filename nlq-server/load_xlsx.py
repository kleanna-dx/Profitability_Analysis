#!/usr/bin/env python3
"""
2.xlsx -> bw_profitability_data 스트림 적재 스크립트
- openpyxl read_only 모드로 메모리 최소화
- 200행씩 배치 INSERT
- 기존 데이터 전체 삭제 후 INSERT
"""
import time
import pymysql
from openpyxl import load_workbook

XLSX_PATH = '/home/user/uploaded_files/2.xlsx'
BATCH_SIZE = 200

def main():
    print("[적재] 2.xlsx -> bw_profitability_data 시작")
    total_start = time.time()

    conn = pymysql.connect(
        host='localhost', user='company', password='company1234!',
        database='company_board', charset='utf8mb4',
        autocommit=False
    )
    cursor = conn.cursor()

    # DB 컬럼 목록 (SEQ 제외)
    cursor.execute("""
        SELECT COLUMN_NAME FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA='company_board' AND TABLE_NAME='bw_profitability_data'
        AND COLUMN_NAME != 'SEQ'
        ORDER BY ORDINAL_POSITION
    """)
    db_col_set = set(row[0] for row in cursor.fetchall())
    print(f"[적재] DB 컬럼: {len(db_col_set)}개 (SEQ 제외)")

    # 엑셀 읽기 (read_only 모드)
    print("[적재] 엑셀 파일 열기 (read_only)...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    print(f"[적재] 시트: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")

    row_iter = ws.iter_rows(values_only=True)

    # 1행: 한글 주석 (스킵)
    next(row_iter)

    # 2행: 영문 컬럼명 -> DB 매핑
    excel_cols = [str(c or '').strip() for c in next(row_iter)]
    col_map = []
    for i, col in enumerate(excel_cols):
        if col and col in db_col_set:
            col_map.append((i, col))

    print(f"[적재] 엑셀 컬럼 {len(excel_cols)}개 중 {len(col_map)}개 DB 매핑")

    if len(col_map) == 0:
        print("[적재] 매핑된 컬럼이 없습니다! 중단.")
        wb.close()
        conn.close()
        return

    col_list = ', '.join(f'`{c[1]}`' for c in col_map)
    placeholders = ', '.join(['%s'] * len(col_map))
    insert_sql = f"INSERT INTO bw_profitability_data ({col_list}) VALUES ({placeholders})"

    # 기존 데이터 삭제
    cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
    old_count = cursor.fetchone()[0]
    if old_count > 0:
        print(f"[적재] 기존 데이터 {old_count:,}행 삭제 중...")
        cursor.execute("TRUNCATE TABLE bw_profitability_data")
        conn.commit()
        print("[적재] 기존 데이터 삭제 완료")
    else:
        print("[적재] 기존 데이터 없음")

    # 3행~: 데이터 적재
    print("[적재] 데이터 적재 시작...")
    insert_start = time.time()
    inserted = 0
    errors = 0
    total_data_rows = 0
    batch = []

    for row in row_iter:
        row_list = list(row)

        has_value = any(
            row_list[idx] is not None and str(row_list[idx]).strip() != ''
            for idx, _ in col_map
        )
        if not has_value:
            continue

        total_data_rows += 1

        values = []
        for idx, col_name in col_map:
            v = row_list[idx] if idx < len(row_list) else None
            if v is None or (isinstance(v, str) and v.strip() == ''):
                values.append(None)
            else:
                values.append(v)

        batch.append(tuple(values))

        if len(batch) >= BATCH_SIZE:
            try:
                cursor.executemany(insert_sql, batch)
                conn.commit()
                inserted += len(batch)
            except Exception as e:
                conn.rollback()
                for row_vals in batch:
                    try:
                        cursor.execute(insert_sql, row_vals)
                        conn.commit()
                        inserted += 1
                    except Exception as row_err:
                        conn.rollback()
                        errors += 1
                        if errors <= 10:
                            print(f"  [오류] 행 {total_data_rows}: {str(row_err)[:150]}")
            batch = []

            if total_data_rows % 10000 < BATCH_SIZE:
                elapsed = time.time() - insert_start
                rate = inserted / elapsed if elapsed > 0 else 0
                print(f"  [진행] {total_data_rows:,}행 읽음 / {inserted:,}행 INSERT ({rate:.0f}행/초)")

    # 남은 배치 처리
    if batch:
        try:
            cursor.executemany(insert_sql, batch)
            conn.commit()
            inserted += len(batch)
        except Exception as e:
            conn.rollback()
            for row_vals in batch:
                try:
                    cursor.execute(insert_sql, row_vals)
                    conn.commit()
                    inserted += 1
                except:
                    conn.rollback()
                    errors += 1

    insert_elapsed = time.time() - insert_start
    print(f"[적재] INSERT 소요시간: {insert_elapsed:.1f}초")

    wb.close()

    cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
    final_count = cursor.fetchone()[0]

    print()
    print("=== 적재 결과 ===")
    print(f"  기존 데이터: {old_count:,}행 삭제")
    print(f"  엑셀 데이터행: {total_data_rows:,}행")
    print(f"  INSERT 성공: {inserted:,}행")
    print(f"  INSERT 실패: {errors}건")
    print(f"  DB 총 행수: {final_count:,}행")
    print(f"  전체 소요시간: {time.time()-total_start:.1f}초")

    cursor.close()
    conn.close()

if __name__ == '__main__':
    main()
