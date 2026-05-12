#!/usr/bin/env python3
"""
2.xlsx -> bw_profitability_data 이어서 적재 (resume)
- 이전 적재에서 OOM으로 중단된 분량을 이어서 INSERT
- openpyxl read_only + 메모리 절약 (gc.collect 추가)
- SKIP_ROWS: 이미 적재된 행 수 (3행 이후 데이터행 기준)
"""
import time
import gc
import pymysql
from openpyxl import load_workbook

XLSX_PATH = '/home/user/uploaded_files/2.xlsx'
BATCH_SIZE = 500
SKIP_ROWS = 63344  # 이미 적재된 행 수

def main():
    print(f"[Resume] 2.xlsx -> bw_profitability_data 이어서 적재")
    print(f"[Resume] {SKIP_ROWS:,}행 스킵, 그 이후부터 INSERT")
    total_start = time.time()

    conn = pymysql.connect(
        host='localhost', user='company', password='company1234!',
        database='company_board', charset='utf8mb4',
        autocommit=True
    )
    cursor = conn.cursor()

    # DB 컬럼 목록 (SEQ 제외)
    cursor.execute("""
        SELECT COLUMN_NAME FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA='company_board' AND TABLE_NAME='bw_profitability_data'
        AND COLUMN_NAME != 'SEQ'
        ORDER BY ORDINAL_POSITION
    """)
    db_col_set = set(row[0] for row in cursor.fetchall())
    print(f"[Resume] DB 컬럼: {len(db_col_set)}개 (SEQ 제외)")

    # 엑셀 읽기
    print("[Resume] 엑셀 파일 열기...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    print(f"[Resume] 시트: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")

    row_iter = ws.iter_rows(values_only=True)
    next(row_iter)  # 1행 한글 주석 스킵

    # 2행: 영문 컬럼명
    excel_cols = [str(c or '').strip() for c in next(row_iter)]
    col_map = []
    for i, col in enumerate(excel_cols):
        if col and col in db_col_set:
            col_map.append((i, col))
    print(f"[Resume] 매핑: {len(col_map)}/{len(excel_cols)}개 컬럼")

    col_list = ', '.join(f'`{c[1]}`' for c in col_map)
    placeholders = ', '.join(['%s'] * len(col_map))
    insert_sql = f"INSERT INTO bw_profitability_data ({col_list}) VALUES ({placeholders})"

    # 스킵
    print(f"[Resume] {SKIP_ROWS:,}행 스킵 중...")
    skip_start = time.time()
    for i in range(SKIP_ROWS):
        next(row_iter)
        if (i + 1) % 20000 == 0:
            print(f"  스킵 {i+1:,}/{SKIP_ROWS:,}...")
            gc.collect()
    print(f"[Resume] 스킵 완료 ({time.time()-skip_start:.1f}초)")
    gc.collect()

    # 적재 시작
    print("[Resume] 나머지 데이터 INSERT 시작...")
    insert_start = time.time()
    inserted = 0
    errors = 0
    data_rows = 0
    batch = []

    for row in row_iter:
        row_list = list(row)

        # 빈 행 체크
        has_value = False
        for idx, _ in col_map:
            v = row_list[idx] if idx < len(row_list) else None
            if v is not None and str(v).strip() != '':
                has_value = True
                break
        if not has_value:
            continue

        data_rows += 1

        values = []
        for idx, col_name in col_map:
            v = row_list[idx] if idx < len(row_list) else None
            if v is None or (isinstance(v, str) and v.strip() == ''):
                values.append(None)
            else:
                values.append(v)
        batch.append(tuple(values))

        if len(batch) >= BATCH_SIZE:
            try:
                cursor.executemany(insert_sql, batch)
                inserted += len(batch)
            except Exception as e:
                for rv in batch:
                    try:
                        cursor.execute(insert_sql, rv)
                        inserted += 1
                    except:
                        errors += 1
            batch = []

            if data_rows % 5000 < BATCH_SIZE:
                elapsed = time.time() - insert_start
                rate = inserted / elapsed if elapsed > 0 else 0
                print(f"  [진행] +{data_rows:,}행 / +{inserted:,} INSERT ({rate:.0f}행/초)")
                gc.collect()

    # 남은 배치
    if batch:
        try:
            cursor.executemany(insert_sql, batch)
            inserted += len(batch)
        except:
            for rv in batch:
                try:
                    cursor.execute(insert_sql, rv)
                    inserted += 1
                except:
                    errors += 1

    insert_elapsed = time.time() - insert_start
    wb.close()

    # 최종 확인
    cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
    final_count = cursor.fetchone()[0]

    expected_total = SKIP_ROWS + data_rows

    print(f"\n[Resume] INSERT 소요시간: {insert_elapsed:.1f}초")
    print()
    print("=== Resume 적재 결과 ===")
    print(f"  이전 적재: {SKIP_ROWS:,}행")
    print(f"  이번 적재 데이터행: {data_rows:,}행")
    print(f"  이번 INSERT 성공: {inserted:,}행")
    print(f"  INSERT 실패: {errors}건")
    print(f"  DB 총 행수: {final_count:,}행")
    print(f"  예상 총 행수: {expected_total:,}행")
    print(f"  전체 소요시간: {time.time()-total_start:.1f}초")

    if final_count == expected_total:
        print("  ✅ 전체 적재 완료!")
    elif final_count >= 109244:
        print("  ✅ 109,244행 이상 적재 완료!")
    else:
        print(f"  ⚠️ DB {final_count:,}행 vs 예상 {expected_total:,}행")

    cursor.close()
    conn.close()

if __name__ == '__main__':
    main()
