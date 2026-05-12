#!/usr/bin/env python3
"""
엑셀 파일 프리뷰 분석 (openpyxl read_only 모드)
- 1행: 한글 주석 (korName)
- 2행: 영문 컬럼명 (engName) → DB 매핑 기준
- 3행~: 데이터
- /BIC/ 접두사 자동 제거하여 DB 컬럼과 매핑
- 출력: JSON (stdout)
"""
import sys
import json
import pymysql
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "파일 경로가 필요합니다."}))
        sys.exit(1)

    xlsx_path = sys.argv[1]

    try:
        # DB 컬럼 조회
        conn = pymysql.connect(
            host='localhost', user='company', password='company1234!',
            database='company_board', charset='utf8mb4'
        )
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA='company_board' AND TABLE_NAME='bw_profitability_data'
            ORDER BY ORDINAL_POSITION
        """)
        db_cols = set(row[0] for row in cursor.fetchall())
        cursor.close()
        conn.close()

        # 엑셀 읽기
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        sheet_name = ws.title
        max_row = ws.max_row
        max_col = ws.max_column

        row_iter = ws.iter_rows(values_only=True)

        # 1행: 한글 주석
        row1 = next(row_iter)
        kor_headers = [str(c or '') for c in row1]

        # 2행: 영문 컬럼명
        row2 = next(row_iter)
        eng_headers = [str(c or '').strip() for c in row2]

        # 컬럼 매핑 분석
        mapped = []
        excluded = []

        for i, eng in enumerate(eng_headers):
            if not eng:
                excluded.append({
                    "index": i,
                    "korName": kor_headers[i] if i < len(kor_headers) else f"(열{i+1})",
                    "engName": "(빈 컬럼명)",
                    "reason": "영문 컬럼명 없음"
                })
                continue

            # /BIC/ 접두사 제거
            clean_col = eng.replace('/BIC/', '') if eng.startswith('/BIC/') else eng
            clean_upper = clean_col.upper()

            if clean_upper == 'SEQ':
                excluded.append({
                    "index": i,
                    "korName": kor_headers[i] if i < len(kor_headers) else '',
                    "engName": eng,
                    "reason": "PK 자동생성 컬럼"
                })
            elif clean_col in db_cols:
                mapped.append({
                    "index": i,
                    "korName": kor_headers[i] if i < len(kor_headers) else '',
                    "engName": eng,
                    "dbColumn": clean_col
                })
            else:
                excluded.append({
                    "index": i,
                    "korName": kor_headers[i] if i < len(kor_headers) else '',
                    "engName": eng,
                    "reason": "DB 테이블에 존재하지 않는 컬럼"
                })

        # 3행~: max_row 기반 빠른 행 수 추정 + 샘플 5행만 읽기
        # ★ 전체 행 순회하지 않음 — 대용량 파일(수만~수십만 행)에서 수십 초 → 0.1초로 단축
        total_data_rows = max(0, (max_row or 2) - 2)  # 헤더 2행 제거

        sample_rows = []
        for row in row_iter:
            row_list = list(row)
            # 빈 행 체크
            has_value = False
            for m in mapped:
                idx = m["index"]
                v = row_list[idx] if idx < len(row_list) else None
                if v is not None and str(v).strip() != '':
                    has_value = True
                    break
            if not has_value:
                continue

            sample_obj = {}
            for m in mapped:
                idx = m["index"]
                v = row_list[idx] if idx < len(row_list) else None
                sample_obj[m["dbColumn"]] = str(v) if v is not None else None
            sample_rows.append(sample_obj)
            if len(sample_rows) >= 5:
                break  # 샘플 5행만 읽고 종료

        wb.close()

        result = {
            "sheetName": sheet_name,
            "totalRows": total_data_rows,
            "totalExcelCols": len([e for e in eng_headers if e]),
            "mappedCols": mapped,
            "excludedCols": excluded,
            "mappedCount": len(mapped),
            "excludedCount": len(excluded),
            "previewRows": sample_rows,
            "previewColumns": [m["dbColumn"] for m in mapped],
        }

        print(json.dumps(result, ensure_ascii=False))

    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)

if __name__ == '__main__':
    main()
