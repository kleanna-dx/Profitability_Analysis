#!/usr/bin/env python3
"""
31.xlsx -> bw_profitability_data 추가 적재 (INSERT only, 기존 데이터 유지)
- openpyxl read_only + pymysql batch executemany
- autocommit=True (대용량 트랜잭션 방지)
- BATCH_SIZE 300 (메모리 절약)
- DELETE/TRUNCATE 없음!
"""
import time
import gc
import pymysql
from openpyxl import load_workbook

XLSX_PATH = '/home/user/uploaded_files/31.xlsx'
BATCH_SIZE = 300

def main():
    print("[적재] 31.xlsx -> bw_profitability_data 추가 INSERT 시작")
    print("[적재] ⚠️ 기존 데이터 유지 (DELETE/TRUNCATE 없음)")
    total_start = time.time()

    conn = pymysql.connect(
        host='localhost', user='company', password='company1234!',
        database='company_board', charset='utf8mb4',
        autocommit=True
    )
    cursor = conn.cursor()

    # 기존 행수 확인
    cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
    before_count = cursor.fetchone()[0]
    print(f"[적재] 기존 DB 행수: {before_count:,}행 (유지됨)")

    # DB 컬럼 목록 (SEQ 제외)
    cursor.execute("""
        SELECT COLUMN_NAME FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA='company_board' AND TABLE_NAME='bw_profitability_data'
        AND COLUMN_NAME != 'SEQ'
        ORDER BY ORDINAL_POSITION
    """)
    db_col_set = set(row[0] for row in cursor.fetchall())
    print(f"[적재] DB 컬럼: {len(db_col_set)}개 (SEQ 제외)")

    # 엑셀 읽기
    print("[적재] 엑셀 파일 열기 (read_only 모드)...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    print(f"[적재] 시트: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")

    row_iter = ws.iter_rows(values_only=True)
    next(row_iter)  # 1행 한글 주석 스킵

    # 2행: 영문 컬럼명
    excel_cols = [str(c or '').strip() for c in next(row_iter)]
    col_map = []
    for i, col in enumerate(excel_cols):
        if col and col in db_col_set:
            col_map.append((i, col))
    print(f"[적재] 매핑: {len(col_map)}/{len(excel_cols)}개 컬럼")

    col_list = ', '.join(f'`{c[1]}`' for c in col_map)
    placeholders = ', '.join(['%s'] * len(col_map))
    insert_sql = f"INSERT INTO bw_profitability_data ({col_list}) VALUES ({placeholders})"

    # 적재 시작
    print("[적재] 데이터 INSERT 시작...")
    insert_start = time.time()
    inserted = 0
    errors = 0
    total_rows = 0
    batch = []

    for row in row_iter:
        row_list = list(row)

        # 빈 행 체크
        has_value = False
        for idx, _ in col_map:
            v = row_list[idx] if idx < len(row_list) else None
            if v is not None and str(v).strip() != '':
                has_value = True
                break
        if not has_value:
            continue

        total_rows += 1

        values = []
        for idx, col_name in col_map:
            v = row_list[idx] if idx < len(row_list) else None
            if v is None or (isinstance(v, str) and v.strip() == ''):
                values.append(None)
            else:
                values.append(v)
        batch.append(tuple(values))

        if len(batch) >= BATCH_SIZE:
            try:
                cursor.executemany(insert_sql, batch)
                inserted += len(batch)
            except Exception as e:
                for rv in batch:
                    try:
                        cursor.execute(insert_sql, rv)
                        inserted += 1
                    except:
                        errors += 1
            batch = []

            if total_rows % 5000 < BATCH_SIZE:
                gc.collect()
                elapsed = time.time() - insert_start
                rate = inserted / elapsed if elapsed > 0 else 0
                print(f"  [진행] {total_rows:,}행 / {inserted:,} INSERT ({rate:.0f}행/초)")

    # 남은 배치
    if batch:
        try:
            cursor.executemany(insert_sql, batch)
            inserted += len(batch)
        except:
            for rv in batch:
                try:
                    cursor.execute(insert_sql, rv)
                    inserted += 1
                except:
                    errors += 1

    insert_elapsed = time.time() - insert_start
    wb.close()
    gc.collect()

    # 최종 확인
    cursor.execute("SELECT COUNT(*) FROM bw_profitability_data")
    after_count = cursor.fetchone()[0]

    # CALMONTH별 분포
    cursor.execute("SELECT CALMONTH, COUNT(*) AS cnt FROM bw_profitability_data GROUP BY CALMONTH ORDER BY CALMONTH")
    dist = cursor.fetchall()

    print(f"\n[적재] INSERT 소요시간: {insert_elapsed:.1f}초")
    print()
    print("=" * 40)
    print("  31.xlsx 적재 결과")
    print("=" * 40)
    print(f"  기존 DB 행수:    {before_count:,}행")
    print(f"  엑셀 데이터행:   {total_rows:,}행")
    print(f"  INSERT 성공:     {inserted:,}행")
    print(f"  INSERT 실패:     {errors}건")
    print(f"  현재 DB 총 행수: {after_count:,}행")
    print(f"  예상 총 행수:    {before_count + total_rows:,}행")
    print(f"  소요시간:        {time.time()-total_start:.1f}초")
    print()
    print("  CALMONTH별 분포:")
    for m, c in dist:
        print(f"    {m}: {c:,}행")

    if after_count == before_count + total_rows:
        print(f"\n  ✅ 추가 적재 성공! ({before_count:,} + {total_rows:,} = {after_count:,})")
    else:
        print(f"\n  ⚠️ 차이: DB {after_count:,} vs 예상 {before_count + total_rows:,}")

    cursor.close()
    conn.close()

if __name__ == '__main__':
    main()
